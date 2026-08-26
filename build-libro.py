#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Genera el libro de captura de la línea base.

    python3 build-libro.py

Doce hojas mensuales con los diez indicadores del §13 ya definidos, umbrales
editables en un único sitio, semáforo automático, resumen anual con tendencia y
la hoja de los cinco números del §7. Solo se teclean las casillas amarillas.

Al terminar recalcula con LibreOffice: openpyxl escribe las fórmulas sin
resultado y el libro tiene que salir con sus valores en caché.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import datetime
import pathlib
import re

# El ejercicio estaba tecleado en tres sitios —el nombre del archivo, la
# entradilla y la cabecera de cada hoja mensual—, que son tres oportunidades de
# que uno se quede atrás al cambiar de año.
EJERCICIO = "2026"

RUTA = str(pathlib.Path(__file__).parent / "instrumentos" /
           ("Captura-Linea-Base-Giraldo-%s.xlsx" % EJERCICIO))

# La versión no se teclea en la hoja: se toma del verificador, que es donde
# vive la versión canónica del sistema. Un libro con versión propia se queda
# atrás en la primera publicación y nadie lo nota hasta la Junta.
_v = {}
exec(compile((pathlib.Path(__file__).parent / "version.py").read_text(encoding="utf-8"),
             "version.py", "exec"), _v)
VERSION = _v["VERSION"]

TINTA   = "FF0B1A20"
ENTRADA = Font(name="Arial", size=10, color="FF0000FF")          # azul: dato que se teclea
FORMULA = Font(name="Arial", size=10, color=TINTA)               # negro: calculado
ENLACE  = Font(name="Arial", size=10, color="FF008000")          # verde: viene de otra hoja
NORMAL  = Font(name="Arial", size=10, color=TINTA)
SUAVE   = Font(name="Arial", size=9, color="FF5D7178")
NEGRITA = Font(name="Arial", size=10, bold=True, color=TINTA)
TITULO  = Font(name="Arial", size=14, bold=True, color=TINTA)
CABEZA  = Font(name="Arial", size=9, bold=True, color="FFFFFFFF")

RELLENO_CAB   = PatternFill("solid", fgColor="FF0B1A20")
RELLENO_ENTRA = PatternFill("solid", fgColor="FFFFFF00")         # amarillo: rellénese
RELLENO_ZEBRA = PatternFill("solid", fgColor="FFF2F4F1")
BORDE = Border(*[Side(style="thin", color="FFD5DAD6")] * 4)

MESES = ["01 Enero", "02 Febrero", "03 Marzo", "04 Abril", "05 Mayo", "06 Junio",
         "07 Julio", "08 Agosto", "09 Septiembre", "10 Octubre", "11 Noviembre", "12 Diciembre"]

# nº, nombre, numerador, denominador, formato, sentido, verde, ámbar, umbral publicado
IND = [
 (1,  "Verificaciones regulatorias cerradas", "Verificaciones V1–V11 con evidencia archivada",
      "11 (fijo)", "0%", "mayor", 1.00, 0.80, "11 de 11"),
 (2,  "Bajas voluntarias del equipo", "Ceses a instancia de la persona trabajadora",
      "Plantilla media del mes", "0.0%", "menor", 0.0001, 0.05, "0"),
 (3,  "Reclamaciones de pacientes heredados", "Reclamaciones formales del mes",
      "—", "0", "menor", 0.0001, 1, "0"),
 (4,  "Documentación digitalizada el mismo día", "Actos con historia, consentimiento e imágenes en el día",
      "Actos clínicos del día", "0%", "mayor", 0.98, 0.90, "100 %"),
 (5,  "Producto pendiente", "Importe aceptado y no terminado a fin de mes",
      "—", '#,##0 €', "descendente", 0, 0, "Descendente"),
 (6,  "Tasa de conversión de la primera visita", "PV del mes aceptadas y pagadas en 60 días",
      "Primeras visitas del mes", "0.0%", "mayor", 0.45, 0.35, "Línea base y mejora"),
 (7,  "Pacientes en mantenimiento activo", "Con revisión en 12 meses y próxima cita agendada",
      "Pacientes con implantes del centro", "0.0%", "mayor", 0.60, 0.40, "Creciente"),
 (8,  "Captación por recomendación", "PV cuyo origen declarado es la recomendación",
      "Primeras visitas del mes", "0.0%", "mayor", 0.30, 0.20, "Creciente"),
 (9,  "Ofrecimiento del programa de cuidado", "Cierres con ofrecimiento registrado",
      "Cierres de tratamiento del mes", "0.0%", "mayor", 0.90, 0.80, "90 %"),
 (10, "Auditorías ejecutadas en fecha", "Auditorías realizadas con resultado registrado",
      "Auditorías programadas", "0%", "mayor", 1.00, 0.50, "100 %"),
]
PRIMERA = 6                                   # primera fila de datos en las hojas mensuales

# Fecha del libro: la del ejercicio que documenta, para que el archivo no cambie
# de un día para otro sin que haya cambiado su contenido.
FECHA_LIBRO = datetime.datetime(int(EJERCICIO), 1, 1)

wb = Workbook()

# ----------------------------------------------------------------- instrucciones
h = wb.active
h.title = "Instrucciones"
h.sheet_view.showGridLines = False
h.column_dimensions["A"].width = 3
h.column_dimensions["B"].width = 30
h.column_dimensions["C"].width = 96

def linea(fila, rotulo, texto, fuente=NORMAL):
    h.cell(fila, 2, rotulo).font = NEGRITA
    c = h.cell(fila, 3, texto); c.font = fuente; c.alignment = Alignment(wrap_text=True, vertical="top")
    h.row_dimensions[fila].height = 30

h["B2"] = "Captura de la línea base · Centro de Excelencia Implantológica Giraldo"
h["B2"].font = TITULO
h["B3"] = ("Instrumento del §7 y del §13 de la Tesis de Dirección v%s · Ejercicio %s"
           % (VERSION, EJERCICIO))
h["B3"].font = SUAVE

linea(5, "Para qué sirve", "La Tesis declara que el centro no dispone todavía de sus cinco números ni de serie propia en ninguno "
         "de los diez indicadores. Este libro es el instrumento para conseguirlos: se rellena una hoja al mes y el resumen "
         "anual y el semáforo se calculan solos.")
linea(6, "Qué se teclea", "Solo las celdas con fondo amarillo y cifra azul. Todo lo demás son fórmulas: no se sobrescriben.")
linea(7, "Código de color", "Azul sobre amarillo = dato que introduce el centro · Negro = calculado · Verde = viene de otra hoja.")
linea(8, "La regla de reporte", "Un indicador sin dato se reporta en rojo, no en blanco. El semáforo lo hace automáticamente: "
         "mientras no haya cifra, la casilla dice SIN DATO y cuenta como roja en la revisión de Junta.")
linea(9, "El indicador 6", "Tiene cohorte cerrada: el dato de un mes no es definitivo hasta sesenta días después. Se anota "
         "provisional y se corrige en la revisión siguiente. La columna «Estado del dato» sirve para eso.")
linea(10, "Umbrales", "Están en la hoja Definiciones y son supuestos de trabajo, no objetivos aprobados. Se cambian ahí una "
          "sola vez y todas las hojas mensuales los recogen.")
linea(11, "Los cinco números", "Tienen hoja propia. Son la prioridad: sin ellos cualquier objetivo comercial es una opinión.")
linea(12, "Ejemplo", "La hoja «01 Enero» viene con una fila de ejemplo en gris al pie, para mostrar el formato esperado. "
          "Bórrese antes de usarla.")

h["B14"] = "Hojas del libro"; h["B14"].font = NEGRITA
for i, (rotulo, que) in enumerate([
    ("Definiciones", "El diccionario de los diez indicadores y los umbrales editables"),
    ("01 a 12 · meses", "Una hoja de captura por mes, con semáforo automático"),
    ("Resumen anual", "Los doce meses en una sola rejilla, con tendencia y meses con dato"),
    ("Los cinco números", "Costes fijos, equilibrio, primeras visitas necesarias, producto pendiente y colchón"),
]):
    h.cell(15 + i, 2, rotulo).font = NORMAL
    h.cell(15 + i, 3, que).font = SUAVE

# ----------------------------------------------------------------- definiciones
d = wb.create_sheet("Definiciones")
d.sheet_view.showGridLines = False
CAB = ["#", "Indicador", "Numerador", "Denominador", "Umbral publicado",
       "Sentido", "Verde si", "Ámbar si", "Qué queda fuera"]
ANCHOS = [4, 38, 46, 30, 20, 14, 12, 12, 52]
for j, (t, a) in enumerate(zip(CAB, ANCHOS), start=1):
    c = d.cell(4, j, t); c.font = CABEZA; c.fill = RELLENO_CAB
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    d.column_dimensions[get_column_letter(j)].width = a
d["A1"] = "Diccionario de indicadores"; d["A1"].font = TITULO
d["A2"] = ("Definición operativa acordada. Cambiar aquí un umbral lo cambia en las doce hojas mensuales. "
           "Los umbrales son supuestos de trabajo, no objetivos aprobados por la Junta.")
d["A2"].font = SUAVE
d.row_dimensions[4].height = 32

FUERA = [
 "Nada: una verificación en trámite cuenta como no cerrada",
 "Fin de contrato temporal previsto e incapacidades temporales",
 "Quejas verbales resueltas en el acto y registradas como incidencia de nivel 1",
 "Nada",
 "Presupuestos presentados y no aceptados",
 "Urgencias, segundas opiniones y derivaciones a otro centro por criterio clínico",
 "Pacientes con cita agendada pero sin ninguna asistencia en 12 meses",
 "Nada: el origen «no consta» permanece en el denominador",
 "Nada: se mide el ofrecimiento, no la contratación",
 "Nada: una auditoría sin registro cuenta como no realizada",
]
for i, (n, nombre, num, den, fmt, sentido, verde, ambar, publicado) in enumerate(IND):
    f = 5 + i
    valores = [n, nombre, num, den, publicado, sentido, verde, ambar, FUERA[i]]
    for j, v in enumerate(valores, start=1):
        c = d.cell(f, j, v)
        c.font = ENTRADA if j in (7, 8) else NORMAL
        if j in (7, 8):
            c.fill = RELLENO_ENTRA
            c.number_format = fmt if fmt.endswith("%") else "0.00"
        c.alignment = Alignment(wrap_text=True, vertical="top")
        c.border = BORDE
    d.row_dimensions[f].height = 34
d.cell(16, 1, "Los umbrales de las columnas G y H son los únicos valores editables de esta hoja.").font = SUAVE
d.freeze_panes = "A5"

# ----------------------------------------------------------------- hojas mensuales
CABM = ["#", "Indicador", "Numerador", "Denominador", "Resultado", "Umbral",
        "Estado", "Estado del dato", "Notas"]
ANCHM = [4, 38, 16, 16, 14, 18, 14, 16, 44]

def semaforo(fila, i):
    """Fórmula de semáforo, según el sentido del indicador."""
    n, _, _, _, _, sentido, _, _, _ = IND[i]
    r, verde, ambar = "E%d" % fila, "Definiciones!$G$%d" % (5 + i), "Definiciones!$H$%d" % (5 + i)
    if sentido == "mayor":
        prueba = 'IF({r}>={v},"VERDE",IF({r}>={a},"ÁMBAR","ROJO"))'.format(r=r, v=verde, a=ambar)
    elif sentido == "menor":
        prueba = 'IF({r}<={v},"VERDE",IF({r}<={a},"ÁMBAR","ROJO"))'.format(r=r, v=verde, a=ambar)
    else:                       # descendente: se compara con el mes anterior
        prueba = '"COMPARAR"'
    return '=IF(C{f}="","SIN DATO",{p})'.format(f=fila, p=prueba)

for m, mes in enumerate(MESES):
    s = wb.create_sheet(mes)
    s.sheet_view.showGridLines = False
    s["A1"] = "Captura mensual · %s de %s" % (mes[3:], EJERCICIO)
    s["A1"].font = TITULO
    s["A2"] = "Rellénense únicamente las casillas amarillas. El resultado y el semáforo se calculan solos."
    s["A2"].font = SUAVE
    for j, (t, a) in enumerate(zip(CABM, ANCHM), start=1):
        c = s.cell(4, j, t); c.font = CABEZA; c.fill = RELLENO_CAB
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        s.column_dimensions[get_column_letter(j)].width = a
    s.row_dimensions[4].height = 30

    for i, (n, nombre, num, den, fmt, sentido, verde, ambar, publicado) in enumerate(IND):
        f = PRIMERA + i
        s.cell(f, 1, n).font = NORMAL
        s.cell(f, 2, nombre).font = NORMAL
        for col in (3, 4):
            c = s.cell(f, col)
            c.font = ENTRADA; c.fill = RELLENO_ENTRA; c.number_format = "#,##0.##"
        if den == "11 (fijo)":
            s.cell(f, 4, 11).font = FORMULA
            s.cell(f, 4).fill = PatternFill()
        elif den == "—":
            s.cell(f, 4, "—").font = SUAVE
            s.cell(f, 4).fill = PatternFill()
            s.cell(f, 4).alignment = Alignment(horizontal="center")
        # resultado: cociente, o el propio numerador cuando no hay denominador
        if den == "—":
            s.cell(f, 5, '=IF(C{f}="","",C{f})'.format(f=f))
        else:
            s.cell(f, 5, '=IF(OR(C{f}="",D{f}="",D{f}=0),"",C{f}/D{f})'.format(f=f))
        s.cell(f, 5).font = FORMULA
        s.cell(f, 5).number_format = fmt
        s.cell(f, 6, publicado).font = SUAVE
        s.cell(f, 7, semaforo(f, i)).font = FORMULA
        s.cell(f, 7).alignment = Alignment(horizontal="center")
        s.cell(f, 8).font = ENTRADA
        s.cell(f, 8).fill = RELLENO_ENTRA
        s.cell(f, 9).font = ENTRADA
        s.cell(f, 9).fill = RELLENO_ENTRA
        for j in range(1, 10):
            s.cell(f, j).border = BORDE
            if i % 2 and not s.cell(f, j).fill.fgColor.rgb == "FFFFFF00":
                pass
        s.row_dimensions[f].height = 22

    s.cell(PRIMERA + 6, 8, "Provisional")          # el indicador 6 nace provisional
    s.cell(PRIMERA + 6, 8).font = ENTRADA
    aviso = PRIMERA + len(IND) + 1
    s.cell(aviso, 2, "El indicador 5 se compara con el mes anterior: la casilla dice COMPARAR porque el sentido "
                     "«descendente» exige mirar la serie, no un umbral fijo.").font = SUAVE
    s.cell(aviso + 1, 2, "Un indicador en SIN DATO cuenta como rojo en la revisión de Junta.").font = SUAVE
    s.freeze_panes = "C%d" % PRIMERA

# fila de ejemplo, solo en enero
e = wb["01 Enero"]
fe = PRIMERA + len(IND) + 4
e.cell(fe, 2, "EJEMPLO — bórrese antes de usar la hoja").font = Font(name="Arial", size=9, bold=True, color="FF8A5015")
for j, v in enumerate([6, "Tasa de conversión de la primera visita", 38, 84], start=1):
    c = e.cell(fe + 1, j, v); c.font = SUAVE; c.fill = RELLENO_ZEBRA
e.cell(fe + 1, 5, '=IF(OR(C{f}="",D{f}=""),"",C{f}/D{f})'.format(f=fe + 1)).font = SUAVE
e.cell(fe + 1, 5).number_format = "0.0%"
e.cell(fe + 1, 9, "38 de 84 primeras visitas aceptadas y pagadas dentro de los 60 días").font = SUAVE

# ----------------------------------------------------------------- resumen anual
r = wb.create_sheet("Resumen anual")
r.sheet_view.showGridLines = False
r["A1"] = "Resumen anual · los doce meses en una rejilla"; r["A1"].font = TITULO
r["A2"] = ("Todo se recoge de las hojas mensuales. «Meses con dato» es la medida honesta de cuánto sabemos "
           "realmente: la Junta debe exigir que llegue a doce.")
r["A2"].font = SUAVE
cab = ["#", "Indicador"] + [m[3:6] for m in MESES] + ["Meses con dato", "Primero", "Último", "Tendencia"]
for j, t in enumerate(cab, start=1):
    c = r.cell(4, j, t); c.font = CABEZA; c.fill = RELLENO_CAB
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
r.column_dimensions["A"].width = 4
r.column_dimensions["B"].width = 38
for j in range(3, 15):
    r.column_dimensions[get_column_letter(j)].width = 9
for j in range(15, 19):
    r.column_dimensions[get_column_letter(j)].width = 15
r.row_dimensions[4].height = 30

for i, (n, nombre, num, den, fmt, sentido, verde, ambar, publicado) in enumerate(IND):
    f = 5 + i
    fm = PRIMERA + i
    r.cell(f, 1, n).font = NORMAL
    r.cell(f, 2, nombre).font = NORMAL
    for j, mes in enumerate(MESES):
        c = r.cell(f, 3 + j, "=IF('{m}'!E{fm}=\"\",\"\",'{m}'!E{fm})".format(m=mes, fm=fm))
        c.font = ENLACE; c.number_format = fmt; c.border = BORDE
    rango = "C{f}:N{f}".format(f=f)
    r.cell(f, 15, "=COUNT({r})".format(r=rango)).font = FORMULA
    r.cell(f, 16, '=IFERROR(INDEX({r},MATCH(TRUE,INDEX({r}<>"",0),0)),"")'.format(r=rango)).font = FORMULA
    r.cell(f, 17, '=IFERROR(LOOKUP(9.99E+307,{r}),"")'.format(r=rango)).font = FORMULA
    mejor = ">" if sentido == "mayor" else "<"
    r.cell(f, 18, '=IF(O{f}<2,"—",IF(Q{f}=P{f},"Estable",IF(Q{f}{s}P{f},"Mejora","Empeora")))'
           .format(f=f, s=mejor)).font = FORMULA
    for j in (16, 17):
        r.cell(f, j).number_format = fmt
    for j in range(1, 19):
        r.cell(f, j).border = BORDE
    r.row_dimensions[f].height = 22
r.cell(16, 2, "El indicador 5 mide euros, no porcentaje: su tendencia «Mejora» significa que el producto "
              "pendiente baja.").font = SUAVE
r.freeze_panes = "C5"

# ----------------------------------------------------------------- los cinco números
n5 = wb.create_sheet("Los cinco números")
n5.sheet_view.showGridLines = False
n5["A1"] = "Los cinco números que aún no tenemos"; n5["A1"].font = TITULO
n5["A2"] = ("§7 de la Tesis. Rellénense las casillas amarillas; los cinco resultados se calculan solos. "
            "Hasta que estén, cualquier objetivo comercial es una opinión.")
n5["A2"].font = SUAVE
n5.column_dimensions["A"].width = 4
n5.column_dimensions["B"].width = 46
n5.column_dimensions["C"].width = 18
n5.column_dimensions["D"].width = 14
n5.column_dimensions["E"].width = 58

ENTRADAS = [
 ("Costes fijos mensuales", "", "#,##0 €", "Nóminas, alquiler, suministros, seguros, cuotas y amortizaciones. Día 15."),
 ("Margen de contribución sobre venta", 0.40, "0%", "Supuesto de trabajo del §18 mientras no haya escandallo propio."),
 ("Ticket medio del caso aceptado", 1800, "#,##0 €", "Supuesto de trabajo del §18. Se sustituye por el real."),
 ("Tasa de conversión de la primera visita", 0.45, "0.0%", "Enlaza con el indicador 6 en cuanto haya tres meses de serie."),
 ("Días laborables al mes", 21, "0", "Parámetro de capacidad."),
 ("Tesorería disponible", "", "#,##0 €", "Saldo libre de compromisos a fecha de corte. Día 21."),
 ("Producto pendiente heredado", "", "#,##0 €", "Importe cobrado o comprometido y no ejecutado. Día 15."),
]
n5.cell(4, 2, "Entradas").font = NEGRITA
for i, (rotulo, valor, fmt, nota) in enumerate(ENTRADAS):
    f = 5 + i
    n5.cell(f, 2, rotulo).font = NORMAL
    c = n5.cell(f, 3, valor if valor != "" else None)
    c.font = ENTRADA; c.fill = RELLENO_ENTRA; c.number_format = fmt; c.border = BORDE
    n5.cell(f, 5, nota).font = SUAVE

n5.cell(14, 2, "Resultados").font = NEGRITA
RES = [
 ("1 · Costes fijos mensuales", '=IF(C5="","pendiente",C5)', "#,##0 €",
  "El primero de los cinco. Sin él no hay ninguno de los demás."),
 ("2 · Punto de equilibrio mensual", '=IF(OR(C5="",C6=0),"pendiente",C5/C6)', "#,##0 €",
  "Facturación mensual necesaria para cubrir los costes fijos, dado el margen de contribución."),
 ("3 · Primeras visitas necesarias al día", '=IF(OR(C15="",C16="pendiente",C7=0,C8=0,C9=0),"pendiente",C16/(C8*C7)/C9)', "0.0",
  "Cuántas primeras visitas diarias hacen falta para llegar al equilibrio con la conversión y el ticket actuales."),
 ("4 · Producto pendiente heredado", '=IF(C11="","pendiente",C11)', "#,##0 €",
  "Caja ya cobrada que hay que convertir en producción. Es la palanca 1 del §14."),
 ("5 · Meses de colchón de tesorería", '=IF(OR(C10="",C5="",C5=0),"pendiente",C10/C5)', "0.0",
  "Cuántos meses aguanta el centro sin ingresar nada. Por debajo de tres, es un riesgo de Junta."),
]
for i, (rotulo, formula, fmt, nota) in enumerate(RES):
    f = 15 + i
    n5.cell(f, 2, rotulo).font = NEGRITA
    c = n5.cell(f, 3, formula)
    c.font = FORMULA; c.number_format = fmt; c.border = BORDE
    n5.cell(f, 5, nota).font = SUAVE
    n5.row_dimensions[f].height = 20
n5.cell(21, 2, "Los resultados dicen «pendiente» mientras falte alguna entrada. Es deliberado: "
               "un número inventado es peor que un hueco declarado.").font = SUAVE

# Las propiedades del libro no pueden ser las que pone openpyxl por su cuenta:
# el autor salía como «openpyxl», el título vacío, y la fecha de creación era la
# del momento de construir, de modo que dos construcciones del mismo contenido
# daban dos archivos distintos y el repositorio quedaba sucio después de cada
# una. Se fijan aquí, con la fecha de la versión y no la del reloj.
wb.properties.creator = "Centro de Excelencia Implantológica Giraldo"
wb.properties.lastModifiedBy = "Centro de Excelencia Implantológica Giraldo"
wb.properties.title = "Captura de la línea base · %s" % EJERCICIO
wb.properties.subject = "Los diez indicadores del §13 de la Tesis de Dirección"
wb.properties.description = (
    "Instrumento del §7 y del §13. Uso interno. Una hoja por mes, resumen anual "
    "y los cinco números. Versión v%s del sistema documental." % VERSION)
wb.properties.category = "Uso interno y confidencial"
wb.properties.language = "es-ES"
wb.properties.revision = None
wb.properties.created = FECHA_LIBRO
wb.properties.modified = FECHA_LIBRO

pathlib.Path(RUTA).parent.mkdir(parents=True, exist_ok=True)
wb.save(RUTA)

# Recalcular es parte de generar, no un paso aparte que alguien recuerde: un
# libro publicado con las fórmulas sin resultado enseña casillas vacías a
# quien lo abra con un visor que no calcule.
from recalc import recalcula
recalcula(RUTA)


def normaliza(ruta, cuando):
    """Deja el libro idéntico byte a byte entre dos construcciones iguales.

    Un .xlsx es un zip, y tanto openpyxl como LibreOffice escriben dentro la
    hora del reloj: la de creación en docProps/core.xml y la de cada entrada del
    zip. El resultado era que construir dos veces el mismo contenido daba dos
    archivos distintos, el repositorio quedaba sucio después de cada
    construcción y no había manera de saber, mirando el archivo, si había
    cambiado algo de verdad. Aquí se sustituyen todas esas horas por la del
    ejercicio que el libro documenta, que no depende de cuándo se construya.
    """
    import re as _re, shutil, tempfile, zipfile
    marca = cuando.strftime("%Y-%m-%dT%H:%M:%SZ")
    fecha = (cuando.year, cuando.month, cuando.day,
             cuando.hour, cuando.minute, cuando.second)
    with zipfile.ZipFile(ruta) as zin:
        piezas = [(i, zin.read(i.filename)) for i in zin.infolist()]
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.close()
    with zipfile.ZipFile(tmp.name, "w", zipfile.ZIP_DEFLATED) as zout:
        for info, datos in piezas:
            if info.filename == "docProps/core.xml":
                texto = datos.decode("utf-8")
                texto = _re.sub(r"(<dcterms:(?:created|modified)[^>]*>)[^<]*(</)",
                                r"\g<1>" + marca + r"\g<2>", texto)
                datos = texto.encode("utf-8")
            nuevo = zipfile.ZipInfo(info.filename, date_time=fecha)
            nuevo.compress_type = info.compress_type
            nuevo.external_attr = info.external_attr
            zout.writestr(nuevo, datos)
    shutil.move(tmp.name, ruta)


normaliza(RUTA, FECHA_LIBRO)
print("  → %s" % RUTA)
